# Підключаємо бібліотеку для роботи із файлами
from os import listdir
#Підключаємо бібліотеку для роботи з ексель файлами розширення xlsx
from openpyxl import *
#Підключаємо бібліотеку для роботи з ексель файлами розширення xls
from xlrd import *

#Створюємо та відкриваємо нову книгу та робочий лист із назвою "New_codes_uktzed" де зберігатимемо потрібні нам дані
wb = Workbook()
ws = wb.create_sheet("New_codes_uktzed", 0)
New_codes_uktzed = wb["New_codes_uktzed"]
#Створюємо перший рядок з назвами колонок
New_codes_uktzed.cell(row=1, column=1, value="Код товара")
New_codes_uktzed.cell(row=1, column=2, value="Артикул")
New_codes_uktzed.cell(row=1, column=3, value="Назва товара згідно інвойсу")
New_codes_uktzed.cell(row=1, column=4, value="Код товара УКТЗЕД")
#Встановлюємо ширину кожної колонки
New_codes_uktzed.column_dimensions['A'].width = 11
New_codes_uktzed.column_dimensions['B'].width = 30
New_codes_uktzed.column_dimensions['C'].width = 27
New_codes_uktzed.column_dimensions['D'].width = 19

#Дізнаємося з якими файлами будемо працювати та записуємо їх у список
allFiles = listdir(".")
allXlsFilesInCurrentFolder = [] #Список всіх файлів xls з новими кодами УКТЗЕД
for item in allFiles:
    pos = item.find(".")
    filenameExtension = item[(pos + 1):]
    if filenameExtension == "xls":
        allXlsFilesInCurrentFolder.append(item)

#Відкриваємо кожен файл із списку allXlsFilesInCurrentFolder, зчитуємо дані та записуємо у новий файл
numberOfRowRead = 5 #Номер рядку з якого починаємо записувати дані
numberOfRowWrite = 2 #Номер рядку з якого починаємо записувати дані
for item in allXlsFilesInCurrentFolder:
        #Відкриваємо перший файл
        temp = open_workbook(item)
        sh = temp.sheet_by_index(0)
        #Циклом проходився по всіх рядках відкритого файла
        for doc in range(5, sh.nrows):
                #Код товара
                code = sh.cell_value(doc, 0)
                New_codes_uktzed.cell(numberOfRowWrite, column=1, value=code)
                #Артикул
                art = sh.cell_value(doc, 1)
                New_codes_uktzed.cell(numberOfRowWrite, column=2, value=art)
                #Назва товара згідно інвойсу
                name = sh.cell_value(doc, 2)
                New_codes_uktzed.cell(numberOfRowWrite, column=3, value=name)
                #Код УКТЗЕД
                uktzed = sh.cell_value(doc, 10)
                New_codes_uktzed.cell(numberOfRowWrite, column=4, value=uktzed)
                #Збільшуємо лічильник рядків для записуємого файла
                numberOfRowWrite = numberOfRowWrite + 1

#Зберігаємо ексель книгу
wb.save('All_new_codes_uktzed.xlsx')
